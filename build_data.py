import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("/home/claude/synthetic_data.json", encoding="utf-8") as f:
    data = json.load(f)

wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=13, color="1F4E78")
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="666666")
BODY_FONT = Font(name="Arial", size=10)

col_order = ["month", "sales_forecast_request", "current_supply_capacity",
             "lead_time_weeks", "moq", "unit_cost_eur", "inventory_on_hand"]
col_labels = ["Month", "Sales Forecast Request", "Current Supply Capacity",
              "Lead Time (weeks)", "MOQ", "Unit Cost (EUR)", "Inventory On Hand"]

# README / Overview sheet
ov = wb.create_sheet("README")
ov["A1"] = "AI Multi-Agent S&OP Platform — Synthetic Data"
ov["A1"].font = TITLE_FONT
ov["A3"] = data["meta"]["company_name"]
ov["A3"].font = BODY_FONT
ov["A4"] = data["meta"]["disclaimer"]
ov["A4"].font = NOTE_FONT
ov["A6"] = "Field Definitions"
ov["A6"].font = Font(name="Arial", bold=True, size=11)
r = 7
for k, v in data["meta"]["field_definitions"].items():
    ov[f"A{r}"] = k
    ov[f"A{r}"].font = Font(name="Arial", bold=True, size=9)
    ov[f"B{r}"] = v
    ov[f"B{r}"].font = BODY_FONT
    r += 1
ov.column_dimensions["A"].width = 26
ov.column_dimensions["B"].width = 60

for key, sc in data["scenarios"].items():
    sh = wb.create_sheet(sc["title"][:31])
    sh["A1"] = sc["title"]
    sh["A1"].font = TITLE_FONT
    sh.merge_cells("A1:L1")
    lifecycle_note = f"    Lifecycle: {sc['product_lifecycle_stage']}" + (
        f" (launch: {sc['launch_month']})" if sc.get("launch_month") else "")
    sh["A2"] = f"Product: {sc['product']}    Market: {sc['market']}    Unit: {sc['unit']}{lifecycle_note}"
    sh["A2"].font = BODY_FONT
    sh.merge_cells("A2:L2")
    sh["A3"] = sc["narrative_context"]
    sh["A3"].font = NOTE_FONT
    sh["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    sh.merge_cells("A3:L3")
    sh.row_dimensions[3].height = 60

    is_module = "powerclass_wp" in sc
    header_row = 5

    if is_module:
        labels = ["Month", "Sales Forecast (qty)", "Supply Capacity (qty)",
                  "Lead Time (weeks)", "MOQ (qty)", "Unit Cost (EUR/module)",
                  "Inventory On Hand (qty)", "Power Class (Wp)",
                  "Sales Forecast (MWp)", "Supply Capacity (MWp)",
                  "Target Inventory (1.5mo, qty)", "Inventory Gap (qty)"]
        for i, label in enumerate(labels, start=1):
            c = sh.cell(row=header_row, column=i, value=label)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        for r_i, row in enumerate(sc["rows"], start=header_row + 1):
            sh.cell(row=r_i, column=1, value=row["month"]).font = BODY_FONT
            sh.cell(row=r_i, column=2, value=row["sales_forecast_request"]).font = BODY_FONT
            sh.cell(row=r_i, column=3, value=row["current_supply_capacity"]).font = BODY_FONT
            sh.cell(row=r_i, column=4, value=row["lead_time_weeks"]).font = BODY_FONT
            sh.cell(row=r_i, column=5, value=row["moq"]).font = BODY_FONT
            cost_cell = sh.cell(row=r_i, column=6, value=row["unit_cost_eur"])
            cost_cell.font = BODY_FONT
            cost_cell.number_format = '€#,##0'
            sh.cell(row=r_i, column=7, value=row["inventory_on_hand"]).font = BODY_FONT
            pc_cell = sh.cell(row=r_i, column=8, value=sc["powerclass_wp"])
            pc_cell.font = Font(name="Arial", size=10, color="0000FF")  # blue = hardcoded input
            # Live formulas, not hardcoded values: MWp = POWERCLASS(Wp) x quantity / 10^6
            f_col_sales = sh.cell(row=r_i, column=9, value=f"=H{r_i}*B{r_i}/1000000")
            f_col_sales.font = BODY_FONT
            f_col_sales.number_format = '0.000'
            f_col_supply = sh.cell(row=r_i, column=10, value=f"=H{r_i}*C{r_i}/1000000")
            f_col_supply.font = BODY_FONT
            f_col_supply.number_format = '0.000'
            # Target inventory = 1.5 x sales forecast (qty); Gap = on-hand - target
            f_target = sh.cell(row=r_i, column=11, value=f"=1.5*B{r_i}")
            f_target.font = BODY_FONT
            f_gap = sh.cell(row=r_i, column=12, value=f"=G{r_i}-K{r_i}")
            f_gap.font = BODY_FONT
        for i in range(1, len(labels) + 1):
            sh.column_dimensions[get_column_letter(i)].width = 20
    else:
        labels = col_labels + ["Target Inventory (1.5mo)", "Inventory Gap"]
        for i, label in enumerate(labels, start=1):
            c = sh.cell(row=header_row, column=i, value=label)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        for r_i, row in enumerate(sc["rows"], start=header_row + 1):
            for c_i, field in enumerate(col_order, start=1):
                cell = sh.cell(row=r_i, column=c_i, value=row[field])
                cell.font = BODY_FONT
                if field == "unit_cost_eur":
                    cell.number_format = '€#,##0'
            # Target inventory = 1.5 x sales forecast (column B); Gap = on-hand (col G) - target (col H)
            f_target = sh.cell(row=r_i, column=8, value=f"=1.5*B{r_i}")
            f_target.font = BODY_FONT
            f_gap = sh.cell(row=r_i, column=9, value=f"=G{r_i}-H{r_i}")
            f_gap.font = BODY_FONT
        for i in range(1, len(labels) + 1):
            sh.column_dimensions[get_column_letter(i)].width = 20

wb.save("/home/claude/synthetic_data.xlsx")
print("XLSX written.")
